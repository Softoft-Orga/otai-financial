from __future__ import annotations

from concurrent.futures import ThreadPoolExecutor
from dataclasses import asdict, is_dataclass
from io import BytesIO
from typing import Any

import pandas as pd
import plotly.graph_objects as go
import plotly.io as pio
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def _maybe_asdict(x: Any) -> Any:
    if is_dataclass(x):
        return asdict(x)
    if hasattr(x, "model_dump"):
        return x.model_dump()
    return x


def _assumptions_df(assumptions: Any | None) -> pd.DataFrame | None:
    if assumptions is None:
        return None
    a = _maybe_asdict(assumptions)
    if not isinstance(a, dict):
        return None
    return pd.DataFrame(
        [{"Parameter": k, "Value": v} for k, v in a.items()],
        columns=["Parameter", "Value"],
    )


def _decisions_df(monthly_decisions: Any | None) -> pd.DataFrame | None:
    if monthly_decisions is None:
        return None
    if not isinstance(monthly_decisions, list):
        return None
    rows: list[dict] = []
    for i, d in enumerate(monthly_decisions):
        dd = _maybe_asdict(d)
        if isinstance(dd, dict):
            rows.append({"month": i, **dd})
    return pd.DataFrame(rows) if rows else None


def _kpi_rows(df: pd.DataFrame) -> list[tuple[str, float | int | None]]:
    def first_neg_month(series_month: pd.Series, series_cash: pd.Series):
        m = series_month[series_cash < 0]
        return int(m.iloc[0]) if len(m) else None

    return [
        (
            "Months",
            int(df["month"].max() + 1) if "month" in df.columns and len(df) else None,
        ),
        (
            "Starting cash (€)",
            float(df["cash"].iloc[0]) if "cash" in df.columns and len(df) else None,
        ),
        (
            "Ending cash (€)",
            float(df["cash"].iloc[-1]) if "cash" in df.columns and len(df) else None,
        ),
        (
            "Ending debt (€)",
            float(df["debt"].iloc[-1]) if "debt" in df.columns and len(df) else None,
        ),
        (
            "Min cash (€)",
            float(df["cash"].min()) if "cash" in df.columns and len(df) else None,
        ),
        (
            "Month of min cash",
            int(df.loc[df["cash"].idxmin(), "month"])
            if "cash" in df.columns and len(df)
            else None,
        ),
        (
            "First month cash < 0",
            first_neg_month(df["month"], df["cash"])
            if "cash" in df.columns and len(df)
            else None,
        ),
        (
            "Total revenue (€)",
            float(df["revenue_total"].sum())
            if "revenue_total" in df.columns and len(df)
            else None,
        ),
        (
            "Avg revenue last 3 months (€)",
            float(df["revenue_total"].tail(3).mean())
            if "revenue_total" in df.columns and len(df) >= 3
            else None,
        ),
        (
            "Total costs ex tax (€)",
            float(df["costs_ex_tax"].sum())
            if "costs_ex_tax" in df.columns and len(df)
            else None,
        ),
        (
            "Total tax (€)",
            float(df["tax"].sum()) if "tax" in df.columns and len(df) else None,
        ),
        (
            "Total profit after tax (€)",
            float((df["profit_bt"] - df["tax"]).sum())
            if "profit_bt" in df.columns and "tax" in df.columns and len(df)
            else None,
        ),
        (
            "End Free active",
            float(df["free_active"].iloc[-1])
            if "free_active" in df.columns and len(df)
            else None,
        ),
        (
            "End Pro active",
            float(df["pro_active"].iloc[-1])
            if "pro_active" in df.columns and len(df)
            else None,
        ),
        (
            "End Enterprise active",
            float(df["ent_active"].iloc[-1])
            if "ent_active" in df.columns and len(df)
            else None,
        ),
        (
            "Total Ads spend (€)",
            float(df["ads_spend"].sum())
            if "ads_spend" in df.columns and len(df)
            else None,
        ),
        (
            "Total SEO spend (€)",
            float(df["organic_marketing_spend"].sum())
            if "organic_marketing_spend" in df.columns and len(df)
            else None,
        ),
        (
            "Total Direct outreach spend (€)",
            float(df["direct_candidate_outreach_spend"].sum())
            if "direct_candidate_outreach_spend" in df.columns and len(df)
            else None,
        ),
    ]


def _kpis_df(df: pd.DataFrame) -> pd.DataFrame:
    return pd.DataFrame(_kpi_rows(df), columns=["Metric", "Value"])


def _with_scenario_column(df: pd.DataFrame, scenario_name: str) -> pd.DataFrame:
    out = df.copy()
    out.insert(0, "Scenario", scenario_name)
    return out


def _build_tables(
    df: pd.DataFrame,
    assumptions: Any | None,
    monthly_decisions: Any | None,
) -> dict[str, pd.DataFrame]:
    overview_cols = [
        "month",
        "cash",
        "debt",
        "revenue_total",
        "costs_ex_tax",
        "tax",
        "net_cashflow",
        "website_users",
        "domain_rating",
        "leads_total",
        "new_free",
        "new_pro",
        "new_ent",
        "free_active",
        "pro_active",
        "ent_active",
        "product_value",
    ]

    finance_cols = [
        "month",
        "revenue_pro",
        "revenue_ent",
        "revenue_total",
        "annual_revenue_ttm",
        "interest_payment",
        "new_credit_draw",
        "debt_repayment",
        "costs_ex_tax",
        "profit_bt",
        "tax",
        "net_cashflow",
        "cash",
        "debt",
    ]

    acq_cols = [
        "month",
        "ads_spend",
        "organic_marketing_spend",
        "dev_spend",
        "operating_spend",
        "direct_candidate_outreach_spend",
        "ads_clicks",
        "domain_rating",
        "website_users",
        "website_leads",
        "website_leads_available",
        "new_direct_leads",
        "direct_contacted_leads",
        "new_direct_demo_appointments",
        "direct_demo_appointments_available",
        "leads_total",
    ]

    funnel_cols = [
        "month",
        "new_free",
        "new_pro",
        "new_ent",
        "churned_free",
        "churned_pro",
        "churned_ent",
        "free_active",
        "pro_active",
        "ent_active",
    ]

    product_cols = [
        "month",
        "product_value",
        "pro_price",
        "ent_price",
        "conv_web_to_lead_eff",
        "conv_website_lead_to_free_eff",
        "conv_website_lead_to_pro_eff",
        "conv_website_lead_to_ent_eff",
        "direct_contacted_demo_conversion_eff",
        "direct_demo_appointment_conversion_to_free_eff",
        "direct_demo_appointment_conversion_to_pro_eff",
        "direct_demo_appointment_conversion_to_ent_eff",
        "upgrade_free_to_pro_eff",
        "upgrade_pro_to_ent_eff",
        "churn_pro_eff",
    ]

    def pick(cols: list[str]) -> pd.DataFrame:
        return df[[c for c in cols if c in df.columns]].copy()

    df_a = _assumptions_df(assumptions)
    df_d = _decisions_df(monthly_decisions)

    tables: dict[str, pd.DataFrame] = {
        "Dashboard_KPIs": _kpis_df(df),
        "Dashboard_Overview": pick(overview_cols),
        "Finance": pick(finance_cols),
        "Acquisition": pick(acq_cols),
        "Funnel": pick(funnel_cols),
        "Product": pick(product_cols),
        "Monthly_Full": df.copy(),
    }
    if df_a is not None:
        tables["Assumptions"] = df_a
    if df_d is not None:
        tables["Monthly_Decisions"] = df_d

    return tables


def _write_excel_report(
    *,
    tables: dict[str, pd.DataFrame],
    df_for_plots: pd.DataFrame | None,
    out_path: str,
    plot_scenarios: list[tuple[str, pd.DataFrame]] | None = None,
) -> str:
    # Import all plotting functions
    from .plots import (
        plot_cash_burn_rate,
        plot_cash_debt_spend,
        plot_cash_position,
        plot_conversion_funnel,
        plot_costs_breakdown,
        plot_customer_acquisition_channels,
        plot_enhanced_dashboard,
        plot_financial_health_score,
        plot_growth_insights,
        plot_growth_metrics_heatmap,
        plot_ltv_cac_analysis,
        plot_market_cap,
        plot_net_cashflow,
        plot_product_value,
        plot_results,
        plot_revenue_breakdown,
        plot_revenue_cashflow,
        plot_revenue_split,
        plot_ttm_revenue,
        plot_unit_economics,
        plot_user_growth,
        plot_user_growth_stacked,
    )

    def generate_plot_images(df: pd.DataFrame) -> dict[str, BytesIO]:
        plot_images = {}

        # Basic plots
        plot_images["User_Growth"] = plot_results(df, save_path=None)
        plot_images["User_Growth_2"] = plot_user_growth(df)
        plot_images["Revenue_Cashflow"] = plot_revenue_cashflow(df)
        plot_images["Cash_Position"] = plot_cash_position(df)
        plot_images["Market_Cap"] = plot_market_cap(df)
        plot_images["Product_Value"] = plot_product_value(df)
        plot_images["Net_Cashflow"] = plot_net_cashflow(df)
        plot_images["TTM_Revenue"] = plot_ttm_revenue(df)

        # Enhanced plots
        plot_images["User_Growth_Stacked"] = plot_user_growth_stacked(df)
        plot_images["Revenue_Breakdown"] = plot_revenue_breakdown(df)
        plot_images["Cash_Burn_Rate"] = plot_cash_burn_rate(df)
        plot_images["Conversion_Funnel"] = plot_conversion_funnel(df)
        plot_images["LTV_CAC_Analysis"] = plot_ltv_cac_analysis(df)
        plot_images["Unit_Economics"] = plot_unit_economics(df)
        plot_images["Growth_Heatmap"] = plot_growth_metrics_heatmap(df)
        plot_images["Acquisition_Channels"] = plot_customer_acquisition_channels(df)
        plot_images["Financial_Health"] = plot_financial_health_score(df)

        # Dashboard plots
        plot_images["Enhanced_Dashboard"] = plot_enhanced_dashboard(df, save_path=None)
        plot_images["Growth_Insights"] = plot_growth_insights(df, save_path=None)
        plot_images["Cash_Debt_Spend"] = plot_cash_debt_spend(df)
        plot_images["Costs_Breakdown"] = plot_costs_breakdown(df)
        plot_images["Revenue_Split"] = plot_revenue_split(df)

        def _render(name_fig: tuple[str, go.Figure]) -> tuple[str, BytesIO]:
            name, fig = name_fig
            buf = BytesIO()
            buf.write(pio.to_image(fig, format="png", scale=2))
            buf.seek(0)
            return name, buf

        items = [(n, f) for n, f in plot_images.items() if f is not None]
        image_bytes = {}
        with ThreadPoolExecutor() as pool:
            for result in pool.map(_render, items):
                image_bytes[result[0]] = result[1]
        return image_bytes

    sheet_order = [
        "Dashboard_KPIs",
        "Dashboard_Overview",
        "Overview_Comparison",
        "Finance",
        "Acquisition",
        "Funnel",
        "Product",
        "Assumptions",
        "Monthly_Decisions",
        "Monthly_Full",
    ]

    w = pd.ExcelWriter(out_path, engine="openpyxl")
    for sheet_name in sheet_order:
        df_sheet = tables.get(sheet_name)
        if df_sheet is not None:
            df_sheet.to_excel(w, index=False, sheet_name=sheet_name)

    wb = w.book

    # Define styles
    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    title_font = Font(size=16, bold=True, color="1F2937")
    subtitle_font = Font(size=14, bold=True, color="374151")

    def autosize(ws):
        for i, col in enumerate(
            ws.iter_cols(
                min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column
            ),
            start=1,
        ):
            max_len = 0
            for cell in col:
                v = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(v))
            ws.column_dimensions[get_column_letter(i)].width = min(
                42, max(10, max_len + 2)
            )

    def style_sheet(ws, is_kpi=False):
        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 30 if is_kpi else 24

        # Style headers
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
            cell.border = border

        # Style data cells
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = border
                if is_kpi and cell.column == 1:  # Metric names in KPI sheet
                    cell.font = Font(bold=True, color="374151")

        autosize(ws)

        # Add title to KPI sheet
        if is_kpi:
            ws.insert_rows(1)
            ws.merge_cells("A1:B1")
            ws["A1"] = "OTAI Financial Simulation - Key Performance Indicators"
            ws["A1"].font = title_font
            ws["A1"].alignment = Alignment(horizontal="center")
            ws.row_dimensions[1].height = 40

    # Style all sheets
    for name in wb.sheetnames:
        is_kpi = name == "Dashboard_KPIs"
        style_sheet(wb[name], is_kpi=is_kpi)

        # Add subtitles to sheets
        if name != "Dashboard_KPIs":
            ws = wb[name]
            ws.insert_rows(1)
            ws.merge_cells(f"A1:{get_column_letter(ws.max_column)}1")

            subtitle_map = {
                "Dashboard_Overview": "Business Overview - Key Metrics Summary",
                "Overview_Comparison": "Scenario Comparison - KPI Overview",
                "Finance": "Financial Performance - Revenue, Costs & Cash Flow",
                "Acquisition": "Customer Acquisition - Marketing & Sales Channels",
                "Funnel": "Conversion Funnel - User Journey Analytics",
                "Product": "Product Metrics - Value, Pricing & Conversions",
                "Assumptions": "Model Assumptions - Input Parameters",
                "Monthly_Decisions": "Monthly Decisions - Strategic Choices",
                "Monthly_Full": "Complete Monthly Data - All Variables",
            }

            ws["A1"] = subtitle_map.get(name, f"{name} Report")
            ws["A1"].font = subtitle_font
            ws["A1"].alignment = Alignment(horizontal="center")
            ws.row_dimensions[1].height = 35

    for name in [
        "Dashboard_Overview",
        "Overview_Comparison",
        "Finance",
        "Acquisition",
        "Funnel",
        "Product",
        "Monthly_Decisions",
        "Monthly_Full",
    ]:
        if name in wb.sheetnames:
            ws = wb[name]
            for col_name in [
                "cash",
                "debt",
                "revenue_total",
                "costs_ex_tax",
                "net_cashflow",
                "tax",
                "profit_bt",
                "revenue_pro",
                "revenue_ent",
                "ads_spend",
                "organic_marketing_spend",
                "dev_spend",
                "operating_spend",
                "direct_candidate_outreach_spend",
                "sales_spend",
                "support_spend",
                "interest_payment",
            ]:
                idx = None
                for j in range(1, ws.max_column + 1):
                    if ws.cell(row=1, column=j).value == col_name:
                        idx = j
                        break
                if idx:
                    for r in range(2, ws.max_row + 1):
                        ws.cell(row=r, column=idx).number_format = '#,##0.00" €"'

    if plot_scenarios is None and df_for_plots is None:
        w.close()
        return out_path

    def add_plot_sheets(
        *,
        image_bytes: dict[str, BytesIO],
        scenario_label: str | None,
        sheet_suffix: str,
    ) -> None:
        sheet_names = [
            f"Plots_Basic{sheet_suffix}",
            f"Plots_Enhanced{sheet_suffix}",
            f"Plots_Dashboards{sheet_suffix}",
        ]
        for sheet_name in sheet_names:
            if sheet_name in wb.sheetnames:
                wb.remove(wb[sheet_name])

        base_title = "Basic Visualization Plots"
        title = (
            f"{base_title} - {scenario_label}"
            if scenario_label
            else base_title
        )
        ws_plots_basic = wb.create_sheet(f"Plots_Basic{sheet_suffix}")
        ws_plots_basic["A1"] = title
        ws_plots_basic["A1"].font = title_font
        ws_plots_basic["A1"].alignment = Alignment(horizontal="center")
        ws_plots_basic.merge_cells("A1:H1")
        ws_plots_basic.row_dimensions[1].height = 40

        basic_plot_names = [
            "User_Growth",
            "Revenue_Cashflow",
            "Cash_Position",
            "Market_Cap",
            "Product_Value",
            "Net_Cashflow",
            "TTM_Revenue",
        ]
        for i, plot_name in enumerate(basic_plot_names):
            if plot_name in image_bytes:
                row = (i // 2) * 40 + 3
                col = (i % 2) * 8 + 1

                img = Image(image_bytes[plot_name])
                img.width = 600
                img.height = 350
                ws_plots_basic.add_image(img, f"{get_column_letter(col)}{row}")

                title_cell = f"{get_column_letter(col)}{row - 1}"
                ws_plots_basic[title_cell] = plot_name.replace("_", " ")
                ws_plots_basic[title_cell].font = Font(size=14, bold=True)
                ws_plots_basic[title_cell].alignment = Alignment(horizontal="center")

        base_title = "Enhanced Analytics Plots"
        title = (
            f"{base_title} - {scenario_label}"
            if scenario_label
            else base_title
        )
        ws_plots_enhanced = wb.create_sheet(f"Plots_Enhanced{sheet_suffix}")
        ws_plots_enhanced["A1"] = title
        ws_plots_enhanced["A1"].font = title_font
        ws_plots_enhanced["A1"].alignment = Alignment(horizontal="center")
        ws_plots_enhanced.merge_cells("A1:H1")
        ws_plots_enhanced.row_dimensions[1].height = 40

        enhanced_plot_names = [
            "User_Growth_Stacked",
            "Revenue_Breakdown",
            "Cash_Burn_Rate",
            "Conversion_Funnel",
            "LTV_CAC_Analysis",
            "Unit_Economics",
            "Growth_Heatmap",
            "Acquisition_Channels",
        ]
        for i, plot_name in enumerate(enhanced_plot_names):
            if plot_name in image_bytes:
                row = (i // 2) * 40 + 3
                col = (i % 2) * 8 + 1

                img = Image(image_bytes[plot_name])
                img.width = 600
                img.height = 350
                ws_plots_enhanced.add_image(img, f"{get_column_letter(col)}{row}")

                title_cell = f"{get_column_letter(col)}{row - 1}"
                ws_plots_enhanced[title_cell] = plot_name.replace("_", " ")
                ws_plots_enhanced[title_cell].font = Font(size=14, bold=True)
                ws_plots_enhanced[title_cell].alignment = Alignment(horizontal="center")

        base_title = "Comprehensive Dashboard Views"
        title = (
            f"{base_title} - {scenario_label}"
            if scenario_label
            else base_title
        )
        ws_plots_dash = wb.create_sheet(f"Plots_Dashboards{sheet_suffix}")
        ws_plots_dash["A1"] = title
        ws_plots_dash["A1"].font = title_font
        ws_plots_dash["A1"].alignment = Alignment(horizontal="center")
        ws_plots_dash.merge_cells("A1:H1")
        ws_plots_dash.row_dimensions[1].height = 40

        dashboard_plot_names = [
            "Enhanced_Dashboard",
            "Growth_Insights",
            "Cash_Debt_Spend",
            "Costs_Breakdown",
            "Revenue_Split",
            "Financial_Health",
        ]
        for i, plot_name in enumerate(dashboard_plot_names):
            if plot_name in image_bytes:
                row = (i // 2) * 45 + 3
                col = (i % 2) * 8 + 1

                img = Image(image_bytes[plot_name])
                img.width = 700
                img.height = 450
                ws_plots_dash.add_image(img, f"{get_column_letter(col)}{row}")

                title_cell = f"{get_column_letter(col)}{row - 1}"
                ws_plots_dash[title_cell] = plot_name.replace("_", " ")
                ws_plots_dash[title_cell].font = Font(size=14, bold=True)
                ws_plots_dash[title_cell].alignment = Alignment(horizontal="center")

        for ws_name in sheet_names:
            ws = wb[ws_name]
            for col in ["A", "B", "C", "D", "E", "F", "G", "H"]:
                ws.column_dimensions[col].width = 15

    if plot_scenarios is None:
        plot_scenarios = [("", df_for_plots)] if df_for_plots is not None else []

    all_bufs: list[BytesIO] = []
    for i, (scenario_label, df) in enumerate(plot_scenarios):
        if df is None:
            continue
        image_bytes = generate_plot_images(df)
        sheet_suffix = "" if plot_scenarios == [("", df_for_plots)] else f"_S{i + 1}"
        add_plot_sheets(
            image_bytes=image_bytes,
            scenario_label=scenario_label or None,
            sheet_suffix=sheet_suffix,
        )
        all_bufs.extend(image_bytes.values())

    w.close()
    for buf in all_bufs:
        buf.close()
    return out_path


def export(
    detailed_log: list[dict] | pd.DataFrame,
    out_path: str = "OTAI_Simulation_Report.xlsx",
    assumptions: Any | None = None,
    monthly_decisions: Any | None = None,
) -> str:
    df = (
        detailed_log
        if isinstance(detailed_log, pd.DataFrame)
        else pd.DataFrame(detailed_log)
    )

    tables = _build_tables(df, assumptions, monthly_decisions)

    return _write_excel_report(tables=tables, df_for_plots=df, out_path=out_path)


def export_scenarios(
    scenario_results: list[dict[str, Any]],
    out_path: str = "OTAI_Simulation_Report.xlsx",
) -> str:
    if not scenario_results:
        raise ValueError("scenario_results must contain at least one scenario")

    combined_tables: dict[str, list[pd.DataFrame]] = {}
    comparison_rows: dict[str, dict[str, float | int | None]] = {}

    for scenario in scenario_results:
        scenario_name = scenario["name"]
        df = scenario["df"]
        assumptions = scenario.get("assumptions")
        decisions = scenario.get("decisions")
        tables = _build_tables(df, assumptions, decisions)
        for sheet_name, table in tables.items():
            combined_tables.setdefault(sheet_name, []).append(
                _with_scenario_column(table, scenario_name)
            )
        comparison_rows[scenario_name] = {
            metric: value for metric, value in _kpi_rows(df)
        }

    final_tables: dict[str, pd.DataFrame] = {
        sheet_name: pd.concat(frames, ignore_index=True)
        for sheet_name, frames in combined_tables.items()
    }

    metric_order = [metric for metric, _ in _kpi_rows(scenario_results[0]["df"])]
    comparison_df = pd.DataFrame(
        [
            {
                "Metric": metric,
                **{name: values.get(metric) for name, values in comparison_rows.items()},
            }
            for metric in metric_order
        ]
    )
    final_tables["Overview_Comparison"] = comparison_df

    return _write_excel_report(
        tables=final_tables,
        df_for_plots=None,
        out_path=out_path,
        plot_scenarios=[(scenario["name"], scenario["df"]) for scenario in scenario_results],
    )


def _month_labels(n: int, start_year: int = 2026, start_month: int = 2) -> list[str]:
    import calendar

    return [
        f"{calendar.month_abbr[(start_month + i - 1) % 12 + 1]} {start_year + (start_month + i - 1) // 12}"
        for i in range(n)
    ]


def export_simple_budget(
    df: pd.DataFrame,
    assumptions: Any,
    out_path: str = "OTAI_Simple_Budget.xlsx",
) -> str:
    """Export a simplified Excel: Overview (with plots), Detailed Monthly Plan, Assumptions."""
    from .plots import (
        plot_cash_position,
        plot_market_cap,
        plot_net_cashflow,
        plot_product_value,
        plot_results,
        plot_revenue_cashflow,
        plot_ttm_revenue,
    )

    n = len(df)
    labels = _month_labels(n)
    r = lambda col: df[col].round(0).values  # noqa: E731
    r1 = lambda col: df[col].round(1).values  # noqa: E731

    # --- Sheet 1: Overview (KPIs + simple summary table) ---
    kpis = _kpis_df(df)

    summary = pd.DataFrame(
        {
            "Month": labels,
            "Cash (€)": r("cash"),
            "Debt (€)": r("debt"),
            "Revenue (€)": r("revenue_total"),
            "Net Cashflow (€)": r("net_cashflow"),
        }
    )

    # --- Sheet 2: Detailed Monthly Plan ---
    plan = pd.DataFrame(
        {
            "Month": labels,
            # Revenue breakdown
            "Rev Pro (€)": r("revenue_pro"),
            "Rev Ent (€)": r("revenue_ent"),
            "Rev Support (€)": r("support_subscription_revenue_total"),
            "Rev Renewals (€)": (df["revenue_renewal_pro"] + df["revenue_renewal_ent"]).round(0).values,
            "Rev Total (€)": r("revenue_total"),
            "TTM Revenue (€)": r("revenue_ttm"),
            # Cost breakdown
            "Ads (€)": r("ads_spend"),
            "SEO (€)": r("organic_marketing_spend"),
            "Dev (€)": r("dev_spend"),
            "Operating (€)": r("operating_expenses"),
            "Outreach (€)": r("direct_candidate_outreach_spend"),
            "Sales (€)": r("sales_spend"),
            "Support (€)": r("support_spend"),
            "Total Costs (€)": r("costs_ex_tax"),
            # Financials
            "Profit BT (€)": r("profit_bt"),
            "Tax (€)": r("tax"),
            "Net Cashflow (€)": r("net_cashflow"),
            "Cash (€)": r("cash"),
            "Debt (€)": r("debt"),
            "Interest (€)": r("interest_payment"),
            "Market Cap (€)": r("market_cap"),
            "Product Value": r("product_value"),
            # Users
            "Free Users": r1("free_active"),
            "Pro Users": r1("pro_active"),
            "Ent Users": r1("ent_active"),
            "New Free": r1("new_free"),
            "New Pro": r1("new_pro"),
            "New Ent": r1("new_ent"),
            # Leads & growth
            "Website Leads": r1("website_leads"),
            "Direct Leads": r1("new_direct_leads"),
            "Domain Rating": r1("domain_rating"),
            "Partners": r1("partners_active"),
        }
    )

    # Money columns in plan (columns B through V = indices 2–22)
    plan_money_cols = list(range(2, 23))
    # Count columns (W onward = indices 23–32)
    plan_count_cols = list(range(23, 33))

    # --- Sheet 3: Assumptions ---
    assumptions_table = _assumptions_df(assumptions)

    # --- Write sheets ---
    w = pd.ExcelWriter(out_path, engine="openpyxl")
    kpis.to_excel(w, index=False, sheet_name="Overview")
    plan.to_excel(w, index=False, sheet_name="Detailed Monthly Plan")
    if assumptions_table is not None:
        assumptions_table.to_excel(w, index=False, sheet_name="Assumptions")

    # --- Styling ---
    wb = w.book
    header_fill = PatternFill("solid", fgColor="4A1A6B")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    title_font = Font(size=16, bold=True, color="4A1A6B")
    section_font = Font(size=14, bold=True, color="4A1A6B")

    for ws in wb.worksheets:
        for i, col in enumerate(
            ws.iter_cols(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column),
            start=1,
        ):
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[get_column_letter(i)].width = min(42, max(12, max_len + 2))
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
        ws.freeze_panes = "A2"

    # --- Detailed Monthly Plan formatting ---
    ws_plan = wb["Detailed Monthly Plan"]
    for col_idx in plan_money_cols:
        for row_idx in range(2, ws_plan.max_row + 1):
            ws_plan.cell(row=row_idx, column=col_idx).number_format = '#,##0" €"'
    for col_idx in plan_count_cols:
        for row_idx in range(2, ws_plan.max_row + 1):
            ws_plan.cell(row=row_idx, column=col_idx).number_format = '#,##0.0'

    # Plan title row
    ws_plan.insert_rows(1)
    ws_plan.merge_cells(f"A1:{get_column_letter(ws_plan.max_column)}1")
    ws_plan["A1"] = "Detailed Monthly Plan — Feb 2026 to Jan 2028"
    ws_plan["A1"].font = title_font
    ws_plan["A1"].alignment = Alignment(horizontal="center")
    ws_plan.row_dimensions[1].height = 40

    # --- Overview sheet ---
    ws_ov = wb["Overview"]

    # Title row
    ws_ov.insert_rows(1)
    ws_ov.merge_cells("A1:B1")
    ws_ov["A1"] = "OTAI Conservative Budget Plan (24 Months)"
    ws_ov["A1"].font = title_font
    ws_ov["A1"].alignment = Alignment(horizontal="center")
    ws_ov.row_dimensions[1].height = 40

    # Simple summary table below KPIs
    sum_start = ws_ov.max_row + 3
    ws_ov.merge_cells(f"A{sum_start}:E{sum_start}")
    ws_ov[f"A{sum_start}"] = "Monthly Summary"
    ws_ov[f"A{sum_start}"].font = section_font
    ws_ov[f"A{sum_start}"].alignment = Alignment(horizontal="center")

    hdr_row = sum_start + 1
    for ci, hdr in enumerate(summary.columns, start=1):
        cell = ws_ov.cell(row=hdr_row, column=ci, value=hdr)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
    for ri, row_data in enumerate(summary.itertuples(index=False), start=hdr_row + 1):
        for ci, val in enumerate(row_data, start=1):
            cell = ws_ov.cell(row=ri, column=ci, value=val)
            if ci >= 2:
                cell.number_format = '#,##0" €"'

    # --- Plots on Overview page ---
    plot_fns = {
        "Simulation Overview": lambda: plot_results(df, save_path=None),
        "Revenue & Cashflow": lambda: plot_revenue_cashflow(df),
        "Cash Position": lambda: plot_cash_position(df),
        "Net Cashflow": lambda: plot_net_cashflow(df),
        "Product Value": lambda: plot_product_value(df),
        "TTM Revenue": lambda: plot_ttm_revenue(df),
        "Market Cap": lambda: plot_market_cap(df),
    }

    # Generate all figures first, then render to PNG in parallel
    plot_items = []
    for plot_title, plot_fn in plot_fns.items():
        fig = plot_fn()
        if fig is not None:
            plot_items.append((plot_title, fig))

    def _render_budget(title_fig: tuple[str, go.Figure]) -> tuple[str, BytesIO]:
        title, fig = title_fig
        buf = BytesIO()
        buf.write(pio.to_image(fig, format="png", width=1400, height=800, scale=2))
        buf.seek(0)
        return title, buf

    rendered_plots: list[tuple[str, BytesIO]] = []
    with ThreadPoolExecutor() as pool:
        rendered_plots = list(pool.map(_render_budget, plot_items))

    current_row = ws_ov.max_row + 3
    ws_ov.merge_cells(f"A{current_row}:L{current_row}")
    ws_ov[f"A{current_row}"] = "Key Charts"
    ws_ov[f"A{current_row}"].font = section_font
    ws_ov[f"A{current_row}"].alignment = Alignment(horizontal="center")
    current_row += 2

    for plot_title, buf in rendered_plots:
        ws_ov.cell(row=current_row, column=1, value=plot_title).font = Font(
            size=13, bold=True, color="4A1A6B",
        )
        img = Image(buf)
        img.width = 900
        img.height = 500
        ws_ov.add_image(img, f"A{current_row + 1}")
        current_row += 28

    w.close()
    return out_path
